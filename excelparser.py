import openpyxl
import sys
from pathlib import Path

def analyze_workbook(file_path: Path) -> str:
    """
    Analyzes an Excel workbook (.xlsx) and generates a structured Markdown report.
    This report includes values, formulas, and inferred semantic relationships.
    """
    if not file_path.exists():
        return f"Error: File not found at {file_path}"

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        return f"Error: Could not open file. Make sure it's a valid .xlsx file. Details: {e}"

    markdown_output = [f"# Analyse der Excel-Datei: `{file_path.name}`\n"]

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        markdown_output.append(f"## Arbeitsblatt: `{sheet_name}`\n")

        # Find used data range to avoid processing empty cells
        min_row, max_row = (sheet.min_row, sheet.max_row)
        min_col, max_col = (sheet.min_column, sheet.max_column)

        # Heuristic to find label-value pairs outside of main tables
        # A simple heuristic: check for text in one column followed by data in the next
        labels = {}
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for i in range(len(row) - 1):
                cell_label = row[i]
                cell_value = row[i+1]
                
                # Check if it looks like a label: text ending with ':' and the next cell has data
                if isinstance(cell_label.value, str) and cell_label.value.strip().endswith(':'):
                    if cell_value.value is not None:
                        label_text = cell_label.value.strip()
                        value_text = f"'{cell_value.value}'"
                        if cell_value.data_type == 'f':
                             value_text = f"`{cell_value.value}` (Formel: `{sheet[cell_value.coordinate].value}`)"
                        labels[label_text] = value_text


        if labels:
            markdown_output.append("### Erkannte Schlüssel-Wert-Paare\n")
            for label, value in labels.items():
                markdown_output.append(f"- **{label}** {value}")
            markdown_output.append("\n---\n")


        markdown_output.append("### Zellinhalt (Werte und Formeln)\n")
        
        # Build a markdown table
        header = [""] + [f"Spalte {chr(ord('A') + i)}" for i in range(min_col-1, max_col)]
        markdown_output.append("| " + " | ".join(header) + " |")
        markdown_output.append("|---" * (len(header) +1) + "|")

        for r_idx, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col), start=min_row):
            row_data = [f"**Zeile {r_idx}**"]
            for cell in row:
                content = ""
                if cell.value is not None:
                    # If the cell contains a formula
                    if cell.data_type == 'f':
                        # In openpyxl, loading with data_only=False gives formula, not value
                        # To show both, you'd need to load twice or use other libraries.
                        # Here, we show the formula.
                        content = f"`{cell.value}`"
                    # If it's a regular value
                    else:
                        content = str(cell.value)
                row_data.append(content)
            markdown_output.append("| " + " | ".join(row_data) + " |")
            
        markdown_output.append("\n")

    return "\n".join(markdown_output)

def main():
    """Main function to run the script from the command line."""
    if len(sys.argv) != 2:
        print("Verwendung: python excel_parser.py <pfad_zur_excel_datei.xlsx>")
        sys.exit(1)

    input_file = Path(sys.argv[1])
    
    # Generate the markdown content
    markdown_content = analyze_workbook(input_file)
    
    # Create the output file name
    output_file = input_file.with_suffix('.md')
    
    # Write the content to the output file
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        print(f"Analyse erfolgreich abgeschlossen. Bericht wurde in `{output_file}` gespeichert.")
    except Exception as e:
        print(f"Fehler beim Schreiben der Datei: {e}")

if __name__ == '__main__':
    main()